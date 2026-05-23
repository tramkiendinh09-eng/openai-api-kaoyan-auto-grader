#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import io
import json
import os
import re
import subprocess
import sys
import threading
import time
import uuid
from dataclasses import dataclass, field
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "auto_grade_exam.py"
MINERU_DIR = ROOT / "tmp" / "mineru_grading"
UPLOAD_DIR = ROOT / "tmp" / "uploads"
OUTPUT_DIR = ROOT / "output" / "auto_grading"
DEFAULT_POLICY_PATH = ROOT / "config" / "kaoyan_math_grading_policy.md"
UPLOAD_ROLES = {"submission", "question_paper", "reference"}
DEFAULT_USER_AGENT = "Codex Desktop/0.133.0-alpha.1 (Windows 10.0.19045; x86_64) unknown (Codex Desktop; 9.41501)"


TASKS: dict[str, "TaskState"] = {}
TASK_LOCK = threading.Lock()


@dataclass
class TaskState:
    run_id: str
    status: str = "queued"
    created_at: float = field(default_factory=time.time)
    started_at: float | None = None
    finished_at: float | None = None
    returncode: int | None = None
    command_preview: list[str] = field(default_factory=list)
    stdout: str = ""
    stderr: str = ""
    output_dir: str = ""
    error: str | None = None
    batch_index: int = 1
    batch_total: int = 1
    child_run_ids: list[str] = field(default_factory=list)
    current_child_run_id: str | None = None
    batch_results: list[dict[str, str]] = field(default_factory=list)
    process: subprocess.Popen | None = field(default=None, repr=False, compare=False)


def json_response(handler: BaseHTTPRequestHandler, payload: object, status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def text_response(handler: BaseHTTPRequestHandler, text: str, status: int = 200, content_type: str = "text/plain") -> None:
    body = text.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", f"{content_type}; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def binary_response(handler: BaseHTTPRequestHandler, body: bytes, filename: str, content_type: str) -> None:
    safe_ascii_name = re.sub(r"[^A-Za-z0-9_.-]", "_", filename) or "download.bin"
    handler.send_response(HTTPStatus.OK)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Length", str(len(body)))
    handler.send_header("Cache-Control", "no-store")
    handler.send_header(
        "Content-Disposition",
        f"attachment; filename=\"{safe_ascii_name}\"; filename*=UTF-8''{quote(filename)}",
    )
    handler.end_headers()
    handler.wfile.write(body)


def read_request_json(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0") or "0")
    raw = handler.rfile.read(length).decode("utf-8")
    if not raw:
        return {}
    payload = json.loads(raw)
    if not isinstance(payload, dict):
        raise ValueError("request body must be a JSON object")
    return payload


def path_role(path: Path) -> str:
    resolved = path.resolve()
    for role in sorted(UPLOAD_ROLES):
        try:
            resolved.relative_to((UPLOAD_DIR / role).resolve())
            return role
        except ValueError:
            continue
    try:
        resolved.relative_to(MINERU_DIR.resolve())
        return "mineru"
    except ValueError:
        return "other"


def is_deletable_upload_pdf(path: Path) -> bool:
    try:
        resolved = path.resolve()
        resolved.relative_to(UPLOAD_DIR.resolve())
    except ValueError:
        return False
    return resolved.is_file() and resolved.suffix.lower() == ".pdf"


def file_row(path: Path) -> dict[str, object]:
    stat = path.stat()
    group = str(path.parent.relative_to(ROOT)) if path.parent.is_relative_to(ROOT) else str(path.parent)
    return {
        "name": path.name,
        "path": str(path),
        "group": group,
        "role": path_role(path),
        "deletable": is_deletable_upload_pdf(path),
        "size_bytes": stat.st_size,
        "mtime": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat.st_mtime)),
        "mtime_ts": stat.st_mtime,
    }


def list_pdf_files() -> list[dict[str, object]]:
    roots = [UPLOAD_DIR, MINERU_DIR, *extra_search_dirs()]
    rows: list[dict[str, object]] = []
    for base in roots:
        if not base.exists():
            continue
        for path in sorted(base.rglob("*.pdf")):
            rows.append(file_row(path))
    return sorted(rows, key=lambda row: float(row.get("mtime_ts", 0)), reverse=True)


def delete_uploaded_pdfs(paths: list[str]) -> dict[str, object]:
    deleted: list[dict[str, str]] = []
    skipped: list[dict[str, str]] = []
    seen: set[str] = set()
    for raw in paths:
        path_text = str(raw or "").strip()
        if not path_text:
            continue
        try:
            path = Path(path_text).expanduser()
            if not path.is_absolute():
                path = (ROOT / path).resolve()
            else:
                path = path.resolve()
            key = str(path).lower()
            if key in seen:
                continue
            seen.add(key)
            if not is_deletable_upload_pdf(path):
                skipped.append({"path": str(path), "reason": "只能删除项目 tmp/uploads 内的 PDF 副本"})
                continue
            name = path.name
            path.unlink()
            deleted.append({"path": str(path), "name": name})
        except Exception as exc:  # noqa: BLE001
            skipped.append({"path": path_text, "reason": f"{type(exc).__name__}: {exc}"})
    return {"deleted": deleted, "skipped": skipped, "pdf_files": list_pdf_files()}


def extra_search_dirs() -> list[Path]:
    raw = os.getenv("GRADER_EXTRA_SEARCH_DIRS", "")
    if not raw.strip():
        return []
    return [Path(part).expanduser() for part in re.split(r"[;\n]", raw) if part.strip()]


def safe_upload_filename(raw_name: str) -> str:
    name = Path(raw_name or "uploaded.pdf").name
    stem = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_. -]+", "_", Path(name).stem).strip(" ._")
    suffix = Path(name).suffix.lower()
    if suffix != ".pdf":
        suffix = ".pdf"
    return f"{stem or 'uploaded'}{suffix}"


def save_uploaded_pdfs(handler: BaseHTTPRequestHandler, role: str) -> list[dict[str, object]]:
    if role not in UPLOAD_ROLES:
        raise ValueError("bad upload role")
    content_type = handler.headers.get("Content-Type", "")
    match = re.search(r"boundary=(?P<boundary>[^;]+)", content_type)
    if "multipart/form-data" not in content_type or not match:
        raise ValueError("upload must be multipart/form-data")
    boundary = match.group("boundary").strip().strip('"').encode("utf-8")
    length = int(handler.headers.get("Content-Length", "0") or "0")
    if length <= 0:
        raise ValueError("empty upload")
    if length > 80 * 1024 * 1024:
        raise ValueError("PDF 超过 80MB，请先压缩或拆分")
    raw = handler.rfile.read(length)
    marker = b"--" + boundary
    uploaded: list[dict[str, object]] = []
    for part in raw.split(marker):
        if b"filename=" not in part:
            continue
        header, sep, body = part.partition(b"\r\n\r\n")
        if not sep:
            continue
        filename_match = re.search(rb'filename="([^"]*)"', header)
        filename = filename_match.group(1).decode("utf-8", errors="replace") if filename_match else "uploaded.pdf"
        body = body.rstrip(b"\r\n-")
        if not body.startswith(b"%PDF"):
            raise ValueError("上传文件不像 PDF，请确认文件格式")
        role_dir = UPLOAD_DIR / role
        role_dir.mkdir(parents=True, exist_ok=True)
        target_name = f"{time.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}_{safe_upload_filename(filename)}"
        target = role_dir / target_name
        target.write_bytes(body)
        uploaded.append(file_row(target))
    if not uploaded:
        raise ValueError("未找到上传的 PDF 文件")
    return uploaded


def require_local_path(raw: str) -> Path:
    if not raw:
        raise ValueError("missing path")
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = (ROOT / path).resolve()
    else:
        path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(path)
    return path


def sanitize_run_id(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9_.-]", "_", value.strip())
    return value[:80] or f"web_{uuid.uuid4().hex[:8]}"


def payload_submission_paths(payload: dict) -> list[Path]:
    raw_values = payload.get("submission_pdfs")
    rows: list[str] = []
    if isinstance(raw_values, list):
        rows.extend(str(item).strip() for item in raw_values if str(item).strip())
    single = str(payload.get("submission_pdf") or "").strip()
    if single:
        rows.append(single)
    deduped: list[Path] = []
    seen: set[str] = set()
    for row in rows:
        path = require_local_path(row)
        key = str(path).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(path)
    return deduped


def build_grading_command(payload: dict, submission_pdf: Path, run_id: str, candidate_name: str) -> tuple[list[str], dict[str, str]]:
    question_paper_pdf_raw = str(payload.get("question_paper_pdf") or "").strip()
    question_paper_pdf = require_local_path(question_paper_pdf_raw) if question_paper_pdf_raw else None
    reference_pdf_raw = str(payload.get("reference_pdf") or "").strip()
    reference_pdf = require_local_path(reference_pdf_raw) if reference_pdf_raw else None
    api_url = str(payload.get("api_url") or "").strip()
    user_agent = str(payload.get("user_agent") or DEFAULT_USER_AGENT).strip()
    model = str(payload.get("model") or "gpt-5.5").strip()
    api_key = str(payload.get("api_key") or "").strip()
    questions = str(payload.get("questions") or "").strip()
    paper_id = str(payload.get("paper_id") or "local-web-paper").strip()
    parse_only = bool(payload.get("parse_only"))
    api_mode = str(payload.get("api_mode") or "responses").strip()
    use_cache = bool(payload.get("use_cache", True))
    reasoning_effort = str(payload.get("reasoning_effort") or "xhigh").strip()
    objective_reasoning_effort = str(payload.get("objective_reasoning_effort") or "high").strip()
    blank_reasoning_effort = str(payload.get("blank_reasoning_effort") or "high").strip()
    solution_reasoning_effort = str(payload.get("solution_reasoning_effort") or "high").strip()
    layout_scan = bool(payload.get("layout_scan", True))
    parallel_visual_rounds = bool(payload.get("parallel_visual_rounds", True))
    objective_batch_mode = bool(payload.get("objective_batch_mode", True))
    single_review = bool(payload.get("single_review", True))
    use_stream = True
    max_retries = int(payload.get("max_retries") or 3)
    timeout_seconds = int(payload.get("timeout_seconds") or 180)
    max_output_tokens = int(payload.get("max_output_tokens") or 100000)
    concurrency = int(payload.get("concurrency") or 10)
    blank_concurrency = int(payload.get("blank_concurrency") or 3)
    solution_concurrency = int(payload.get("solution_concurrency") or 6)
    reference_is_official = bool(payload.get("reference_is_official"))
    strict_official = bool(payload.get("strict_official"))
    policy_path = require_local_path(str(payload.get("policy_path") or str(DEFAULT_POLICY_PATH)))

    if question_paper_pdf is None:
        raise ValueError("请选择试卷题目 PDF")
    if reference_pdf is None:
        raise ValueError("请选择参考答案 / 评分依据 PDF")

    model_required = not parse_only
    if model_required and not api_url:
        raise ValueError("调用模型时必须填写 API URL")
    if model_required and not api_key:
        raise ValueError("调用模型时必须填写 API Key")

    command = [
        sys.executable,
        str(SCRIPT_PATH),
        "--paper-id",
        paper_id,
        "--candidate-name",
        candidate_name,
        "--api-url",
        api_url,
        "--user-agent",
        user_agent,
        "--model",
        model,
        "--api-mode",
        api_mode,
        "--reasoning-effort",
        reasoning_effort,
        "--objective-reasoning-effort",
        objective_reasoning_effort,
        "--blank-reasoning-effort",
        blank_reasoning_effort,
        "--solution-reasoning-effort",
        solution_reasoning_effort,
        "--max-retries",
        str(max_retries),
        "--timeout-seconds",
        str(timeout_seconds),
        "--max-output-tokens",
        str(max_output_tokens),
        "--concurrency",
        str(concurrency),
        "--blank-concurrency",
        str(blank_concurrency),
        "--solution-concurrency",
        str(solution_concurrency),
        "--run-id",
        run_id,
        "--policy-path",
        str(policy_path),
    ]
    command.extend(["--submission-pdf", str(submission_pdf)])
    command.extend(["--question-paper-pdf", str(question_paper_pdf)])
    command.extend(["--reference-pdf", str(reference_pdf)])
    if questions:
        command.extend(["--questions", questions])
    if parse_only:
        command.append("--parse-only")
    if use_cache:
        command.append("--use-cache")
    else:
        command.append("--no-cache")
    if reference_is_official:
        command.append("--reference-is-official")
    if strict_official:
        command.append("--strict-official")
    if parallel_visual_rounds:
        command.append("--parallel-visual-rounds")
    else:
        command.append("--sequential-visual-rounds")
    if objective_batch_mode:
        command.append("--objective-batch-mode")
    else:
        command.append("--per-question-objective")
    if single_review:
        command.append("--single-review")
    else:
        command.append("--double-review")
    command.append("--stream")
    if layout_scan:
        command.append("--layout-scan")
    else:
        command.append("--no-layout-scan")

    env = os.environ.copy()
    if api_key:
        env["GRADER_API_KEY"] = api_key
    env["PYTHONIOENCODING"] = "utf-8"
    return command, env


def start_grading_task(payload: dict) -> TaskState:
    submission_pdfs = payload_submission_paths(payload)
    if not submission_pdfs:
        raise ValueError("请选择至少一份学生答卷 PDF")
    base_candidate_name = str(payload.get("candidate_name") or "local-candidate").strip()
    parent_run_id = sanitize_run_id(f"web_{time.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}")
    child_run_ids = [
        sanitize_run_id(f"{parent_run_id}_{idx:03d}_{safe_stem_for_run(path)}")
        for idx, path in enumerate(submission_pdfs, start=1)
    ]

    state = TaskState(
        run_id=parent_run_id,
        command_preview=["python", "scripts/auto_grade_exam.py", "...", "--batch", str(len(submission_pdfs))],
        output_dir=str(OUTPUT_DIR),
        batch_total=len(submission_pdfs),
        child_run_ids=child_run_ids,
    )
    with TASK_LOCK:
        TASKS[parent_run_id] = state
    thread = threading.Thread(target=run_batch_command, args=(state, payload, submission_pdfs, child_run_ids, base_candidate_name), daemon=True)
    thread.start()
    return state


def safe_stem_for_run(path: Path) -> str:
    stem = re.sub(r"[^A-Za-z0-9_.-]", "_", path.stem.strip())
    return stem[:28] or uuid.uuid4().hex[:8]


def run_command(state: TaskState, command: list[str], env: dict[str, str]) -> None:
    state.status = "running"
    state.started_at = time.time()
    try:
        process = subprocess.Popen(
            command,
            cwd=str(ROOT),
            env=env,
            text=True,
            encoding="utf-8",
            errors="replace",
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        state.process = process
        stdout, stderr = process.communicate()
        state.returncode = process.returncode
        state.stdout = (stdout or "")[-12000:]
        state.stderr = (stderr or "")[-12000:]
        if state.status == "stopping":
            state.status = "failed"
            state.error = "任务已手动停止"
            return
        state.status = "finished" if process.returncode == 0 else "failed"
    except Exception as exc:  # noqa: BLE001
        state.status = "failed"
        state.error = f"{type(exc).__name__}: {exc}"
    finally:
        state.process = None
        state.finished_at = time.time()


def run_batch_command(
    state: TaskState,
    payload: dict,
    submission_pdfs: list[Path],
    child_run_ids: list[str],
    base_candidate_name: str,
) -> None:
    state.status = "running"
    state.started_at = time.time()
    aggregate_stdout: list[str] = []
    aggregate_stderr: list[str] = []
    try:
        for idx, (submission_pdf, child_run_id) in enumerate(zip(submission_pdfs, child_run_ids), start=1):
            if state.status == "stopping":
                state.error = "批量任务已手动停止"
                break
            state.batch_index = idx
            state.current_child_run_id = child_run_id
            candidate_name = base_candidate_name
            if len(submission_pdfs) > 1:
                candidate_name = f"{base_candidate_name}-{idx:03d}-{submission_pdf.stem}"
            command, env = build_grading_command(payload, submission_pdf, child_run_id, candidate_name)
            state.command_preview = ["python", "scripts/auto_grade_exam.py", "...", "--run-id", child_run_id]
            child_state = TaskState(
                run_id=child_run_id,
                status="running",
                started_at=time.time(),
                command_preview=state.command_preview,
                output_dir=str(OUTPUT_DIR / child_run_id),
                batch_index=idx,
                batch_total=len(submission_pdfs),
            )
            with TASK_LOCK:
                TASKS[child_run_id] = child_state
            try:
                process = subprocess.Popen(
                    command,
                    cwd=str(ROOT),
                    env=env,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                )
                state.process = process
                child_state.process = process
                stdout, stderr = process.communicate()
                child_state.returncode = process.returncode
                child_state.stdout = (stdout or "")[-12000:]
                child_state.stderr = (stderr or "")[-12000:]
                child_state.status = "finished" if process.returncode == 0 else "failed"
                child_state.finished_at = time.time()
                aggregate_stdout.append(f"===== {child_run_id} =====\n{stdout or ''}")
                aggregate_stderr.append(f"===== {child_run_id} =====\n{stderr or ''}")
                state.batch_results.append(
                    {
                        "run_id": child_run_id,
                        "status": child_state.status,
                        "submission_pdf": str(submission_pdf),
                        "output_dir": child_state.output_dir,
                    }
                )
                if process.returncode != 0:
                    state.returncode = process.returncode
            except Exception as exc:  # noqa: BLE001
                child_state.status = "failed"
                child_state.error = f"{type(exc).__name__}: {exc}"
                child_state.finished_at = time.time()
                state.returncode = 1
                state.batch_results.append(
                    {
                        "run_id": child_run_id,
                        "status": "failed",
                        "submission_pdf": str(submission_pdf),
                        "output_dir": child_state.output_dir,
                    }
                )
            finally:
                child_state.process = None
                state.process = None
        state.stdout = "\n".join(aggregate_stdout)[-12000:]
        state.stderr = "\n".join(aggregate_stderr)[-12000:]
        if state.status == "stopping":
            state.status = "failed"
            state.error = state.error or "批量任务已手动停止"
        else:
            failed = [row for row in state.batch_results if row.get("status") != "finished"]
            state.status = "failed" if failed else "finished"
            state.returncode = 1 if failed else 0
    except Exception as exc:  # noqa: BLE001
        state.status = "failed"
        state.error = f"{type(exc).__name__}: {exc}"
        state.returncode = 1
    finally:
        state.current_child_run_id = None
        state.process = None
        state.finished_at = time.time()


def stop_grading_task(run_id: str) -> TaskState:
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", run_id):
        raise ValueError("bad run id")
    with TASK_LOCK:
        state = TASKS.get(run_id)
    if state is None:
        raise ValueError("task not found")
    if state.status not in {"queued", "running"}:
        return state
    state.status = "stopping"
    process = state.process
    if process is not None and process.poll() is None:
        process.terminate()
        try:
            process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            process.kill()
    return state


def load_report(run_id: str, filename: str) -> object:
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", run_id):
        raise ValueError("bad run id")
    allowed = {"grading_report.json", "partial_report.json", "parsed_answers.json", "reference_index.json"}
    if filename not in allowed:
        raise ValueError("bad report filename")
    path = OUTPUT_DIR / run_id / filename
    if not path.exists():
        return {"available": False, "path": str(path)}
    return json.loads(path.read_text(encoding="utf-8"))


def load_score_report(run_id: str) -> dict:
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", run_id):
        raise ValueError("bad run id")
    run_dir = OUTPUT_DIR / run_id
    for filename in ("grading_report.json", "partial_report.json"):
        path = run_dir / filename
        if path.exists():
            report = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(report, dict):
                report["_scorebook_source_file"] = filename
                return report
    raise FileNotFoundError(f"未找到 {run_id} 的 grading_report.json 或 partial_report.json")


def scorebook_run_ids(run_id: str) -> list[str]:
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", run_id):
        raise ValueError("bad run id")
    with TASK_LOCK:
        task = TASKS.get(run_id)
        child_run_ids = list(task.child_run_ids) if task and task.child_run_ids else []
    if child_run_ids:
        return child_run_ids
    direct_dir = OUTPUT_DIR / run_id
    if (direct_dir / "grading_report.json").exists() or (direct_dir / "partial_report.json").exists():
        return [run_id]
    if not OUTPUT_DIR.exists():
        return [run_id]
    prefix = run_id + "_"
    inferred = sorted(
        [
            path.name
            for path in OUTPUT_DIR.iterdir()
            if path.is_dir()
            and path.name.startswith(prefix)
            and ((path / "grading_report.json").exists() or (path / "partial_report.json").exists())
        ]
    )
    return inferred or [run_id]


def text_for_excel(value: object, max_chars: int = 2800) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        text = "\n".join(str(item) for item in value if item is not None)
    elif isinstance(value, dict):
        text = json.dumps(value, ensure_ascii=False, default=str)
    else:
        text = str(value)
    text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", text)
    text = re.sub(r"[ \t]+\n", "\n", text).strip()
    if len(text) > max_chars:
        return text[: max_chars - 1] + "…"
    return text


def export_list(value: object) -> list[object]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def numeric_score(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number, 2)


def score_text(value: object) -> str | float:
    number = numeric_score(value)
    return "需复核" if number is None else number


def round_score(row: dict, key: str) -> str | float:
    round_payload = row.get(key)
    if isinstance(round_payload, dict):
        return score_text(round_payload.get("score"))
    return ""


def sheet_title(raw: str, used: set[str]) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "_", raw).strip() or "得分表"
    title = title[:31]
    candidate = title
    suffix = 2
    while candidate in used:
        marker = f"_{suffix}"
        candidate = f"{title[: 31 - len(marker)]}{marker}"
        suffix += 1
    used.add(candidate)
    return candidate


def candidate_label(report: dict, fallback: str) -> str:
    candidate = report.get("candidate") if isinstance(report.get("candidate"), dict) else {}
    name = str(candidate.get("name") or "").strip()
    return name or fallback


def build_scorebook(run_id: str) -> tuple[bytes, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - local dependency guard
        raise RuntimeError("缺少 openpyxl，无法生成 XLSX 得分表。请先安装 openpyxl。") from exc

    child_ids = scorebook_run_ids(run_id)
    reports: list[tuple[str, dict]] = []
    errors: list[tuple[str, str]] = []
    for child_id in child_ids:
        try:
            reports.append((child_id, load_score_report(child_id)))
        except Exception as exc:  # noqa: BLE001
            errors.append((child_id, f"{type(exc).__name__}: {exc}"))
    if not reports and errors:
        raise FileNotFoundError(errors[0][1])
    if not reports:
        raise FileNotFoundError(f"未找到 {run_id} 的可导出报告")

    wb = Workbook()
    used_titles: set[str] = set()
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    title_fill = PatternFill("solid", fgColor="17324D")
    review_fill = PatternFill("solid", fgColor="FFF1F0")
    ok_fill = PatternFill("solid", fgColor="EFFAF4")
    thin = Side(style="thin", color="D9DEE5")
    border = Border(bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    def style_columns(ws) -> None:
        widths = {
            "A": 8,
            "B": 10,
            "C": 10,
            "D": 8,
            "E": 10,
            "F": 10,
            "G": 10,
            "H": 10,
            "I": 10,
            "J": 14,
            "K": 28,
            "L": 28,
            "M": 28,
            "N": 28,
            "O": 22,
            "P": 26,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def apply_table_style(ws, header_row: int, max_col: int) -> None:
        for cell in ws[header_row]:
            if cell.column <= max_col:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="1F2937")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = wrap
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"

    def write_score_sheet(ws, child_run_id: str, report: dict) -> None:
        summary = report.get("score_summary") if isinstance(report.get("score_summary"), dict) else {}
        paper = report.get("paper") if isinstance(report.get("paper"), dict) else {}
        module_scores = report.get("module_scores") if isinstance(report.get("module_scores"), list) else []
        rows = report.get("question_scores") if isinstance(report.get("question_scores"), list) else []
        total_score = score_text(summary.get("total_score_for_graded_questions"))
        full_score = score_text(paper.get("total_full_score_for_graded_questions"))
        ws["A1"] = "考研数学自动阅卷得分表"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = title_fill
        ws.merge_cells("A1:P1")
        ws["A2"] = "Run ID"
        ws["B2"] = child_run_id
        ws["A3"] = "考生"
        ws["B3"] = candidate_label(report, child_run_id)
        ws["A4"] = "试卷"
        ws["B4"] = paper.get("paper_id") or ""
        ws["D2"] = "总分"
        ws["E2"] = f"{total_score} / {full_score}"
        ws["D3"] = "已评题数"
        ws["E3"] = summary.get("graded_question_count", len(rows))
        ws["D4"] = "需复核"
        ws["E4"] = summary.get("human_review_count", sum(1 for row in rows if isinstance(row, dict) and row.get("needs_human_review")))
        ws["G2"] = "报告类型"
        ws["H2"] = report.get("_scorebook_source_file") or ""
        ws["G3"] = "生成时间"
        ws["H3"] = report.get("generated_at") or ""
        ws["G4"] = "模块得分"
        ws["H4"] = "；".join(
            f"{item.get('module')} {score_text(item.get('score'))}/{score_text(item.get('full_score'))}"
            for item in module_scores
            if isinstance(item, dict)
        )
        for key_cell in ("A2", "A3", "A4", "D2", "D3", "D4", "G2", "G3", "G4"):
            ws[key_cell].font = Font(bold=True, color="374151")
        for row_no in range(2, 5):
            for col_no in range(1, 17):
                ws.cell(row=row_no, column=col_no).alignment = wrap

        headers = [
            "题号",
            "题型",
            "最终得分",
            "满分",
            "第一次评分",
            "第二次评分",
            "第三次评分",
            "需人工复核",
            "置信度",
            "评分引擎",
            "识别作答",
            "主要得分点",
            "主要扣分点/原因",
            "有效步骤",
            "错误或缺失步骤",
            "视觉与证据",
        ]
        header_row = 7
        for col_no, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_no, value=header)
        for row in rows:
            if not isinstance(row, dict):
                continue
            source = row.get("student_answer_source") if isinstance(row.get("student_answer_source"), dict) else {}
            visual_sources = row.get("visual_sources") if isinstance(row.get("visual_sources"), dict) else {}
            final_score = score_text(row.get("final_score"))
            excel_row = [
                text_for_excel(row.get("question_no"), 80),
                text_for_excel(row.get("question_type"), 80),
                final_score,
                score_text(row.get("full_score")),
                round_score(row, "grading_round_1"),
                round_score(row, "grading_round_2"),
                round_score(row, "grading_round_3"),
                "是" if row.get("needs_human_review") else "否",
                numeric_score(row.get("confidence")),
                text_for_excel(row.get("scoring_engine") or row.get("evidence_used"), 180),
                text_for_excel(
                    row.get("recognized_student_answer")
                    or source.get("recognized_student_answer")
                    or source.get("student_answer_ocr"),
                    1400,
                ),
                text_for_excel(row.get("main_earned_points") or row.get("earned_points"), 1800),
                text_for_excel(export_list(row.get("main_deducted_points") or row.get("deducted_points")) + ([row.get("review_reason")] if row.get("review_reason") else []), 1800),
                text_for_excel(row.get("valid_student_steps"), 1800),
                text_for_excel(row.get("wrong_or_missing_steps"), 1800),
                text_for_excel(
                    {
                        "evidence_used": row.get("evidence_used") or source.get("evidence_used"),
                        "visual_reading_summary": row.get("visual_reading_summary") or source.get("visual_reading_summary"),
                        "selection_reason": visual_sources.get("selection_reason"),
                    },
                    1800,
                ),
            ]
            ws.append(excel_row)
            current_row = ws.max_row
            if row.get("needs_human_review"):
                ws.cell(current_row, 8).fill = review_fill
            elif row.get("final_score") is not None:
                ws.cell(current_row, 8).fill = ok_fill
        style_columns(ws)
        apply_table_style(ws, header_row, len(headers))

    def write_summary_sheet(ws) -> None:
        ws.title = sheet_title("汇总", used_titles)
        ws["A1"] = "批量阅卷得分表汇总"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = title_fill
        ws.merge_cells("A1:H1")
        headers = ["序号", "Sheet", "Run ID", "考生", "总分", "满分", "已评题数", "需复核"]
        for col_no, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_no, value=header)
        for idx, (child_id, report) in enumerate(reports, start=1):
            summary = report.get("score_summary") if isinstance(report.get("score_summary"), dict) else {}
            paper = report.get("paper") if isinstance(report.get("paper"), dict) else {}
            title = sheet_title(f"{idx:03d}_{candidate_label(report, child_id)}", used_titles)
            ws.append(
                [
                    idx,
                    title,
                    child_id,
                    candidate_label(report, child_id),
                    score_text(summary.get("total_score_for_graded_questions")),
                    score_text(paper.get("total_full_score_for_graded_questions")),
                    summary.get("graded_question_count", len(report.get("question_scores") or [])),
                    summary.get("human_review_count", ""),
                ]
            )
        for col_no, width in enumerate([8, 24, 42, 30, 10, 10, 12, 10], start=1):
            ws.column_dimensions[get_column_letter(col_no)].width = width
        apply_table_style(ws, 3, len(headers))

    if len(reports) > 1:
        write_summary_sheet(wb.active)
        for idx, (child_id, report) in enumerate(reports, start=1):
            title = wb.active.cell(row=idx + 3, column=2).value or f"{idx:03d}"
            ws = wb.create_sheet(str(title))
            write_score_sheet(ws, child_id, report)
    else:
        child_id, report = reports[0]
        ws = wb.active
        ws.title = sheet_title("得分表", used_titles)
        write_score_sheet(ws, child_id, report)

    if errors:
        ws = wb.create_sheet(sheet_title("导出错误", used_titles))
        ws.append(["Run ID", "错误"])
        for child_id, error in errors:
            ws.append([child_id, error])
        apply_table_style(ws, 1, 2)

    output = io.BytesIO()
    wb.save(output)
    filename = f"得分表_{run_id}.xlsx"
    return output.getvalue(), filename


def load_audit_tail(run_id: str) -> dict:
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", run_id):
        raise ValueError("bad run id")
    path = OUTPUT_DIR / run_id / "audit_log.jsonl"
    if not path.exists():
        return {"available": False, "path": str(path), "events": []}
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()[-240:]
    events = []
    for line in lines:
        try:
            events.append(json.loads(line))
        except json.JSONDecodeError:
            events.append({"raw": line})
    return {"available": True, "path": str(path), "events": events, "progress": audit_progress(events)}


def audit_progress(events: list[dict]) -> dict:
    def as_int(value: object, default: int = 0) -> int:
        try:
            return int(value)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return default

    latest_question = None
    latest_round = None
    latest_event = None
    latest_elapsed = None
    latest_error = None
    layout_scan_status = "双xhigh视觉小格定位：尚未看到日志"
    successful_calls = 0
    failed_calls = 0
    active_calls = 0
    model_tasks: set[str] = set()
    retry_or_degrade_calls = 0
    for event in events:
        if not isinstance(event, dict):
            continue
        name = event.get("event")
        if name == "answer_layout_scan_started":
            layout_scan_status = "双xhigh视觉小格定位：已发起"
        elif name == "answer_layout_scan_finished":
            layout_scan_status = f"双xhigh视觉小格定位：已完成，定位 {event.get('located_question_count', 0)} 题，答案串 {event.get('objective_answer_run_count', 0)} 组"
            if event.get("round_errors"):
                layout_scan_status += "，其中一路失败已兜底"
        elif name == "answer_layout_scan_round_failed":
            layout_scan_status = f"双xhigh视觉小格定位：{event.get('round', '')} 路失败，等待另一路或兜底"
        elif name == "answer_layout_scan_compact_fallback_started":
            layout_scan_status = "双xhigh视觉小格定位：双路超时，已启动紧凑high兜底"
        elif name == "answer_layout_scan_compact_fallback_finished":
            layout_scan_status = f"紧凑high视觉定位：已完成，定位 {event.get('located_question_count', 0)} 题，答案串 {event.get('objective_answer_run_count', 0)} 组"
        elif name == "answer_layout_scan_compact_fallback_failed":
            layout_scan_status = "紧凑high视觉定位：失败，已改用页码启发式"
        elif name == "answer_layout_scan_failed":
            layout_scan_status = "双xhigh视觉小格定位：失败，已改用页码启发式"
        if name in {"visual_evidence_attached", "local_objective_grade"} and event.get("question_no"):
            latest_question = str(event.get("question_no"))
            latest_round = "准备视觉阅卷" if name == "visual_evidence_attached" else "本地客观题判分"
            latest_event = name
        elif name == "shared_objective_answer_line_parsed":
            latest_question = "1-10"
            latest_round = "选择题答案行已拆分"
            latest_event = name
        elif name == "objective_batch_local_short_circuit":
            latest_question = "1-10"
            latest_round = "选择题本地确定判分"
            latest_event = name
        elif name == "question_result_cache_hit":
            latest_question = str(event.get("question_no") or latest_question or "")
            latest_round = "单题结果缓存命中"
            latest_event = name
        elif name == "answer_layout_scan_started":
            latest_question = "整卷"
            latest_round = "双xhigh视觉小格定位已发起"
            latest_event = name
        elif name == "answer_layout_scan_finished":
            latest_question = "整卷"
            latest_round = f"双xhigh视觉小格定位完成，定位 {event.get('located_question_count', 0)} 题"
            latest_event = name
        elif name == "answer_layout_scan_round_failed":
            latest_question = "整卷"
            latest_round = f"双xhigh视觉小格定位 {event.get('round', '')} 路失败"
            latest_event = name
            latest_error = str(event.get("error") or "")
        elif name == "answer_layout_scan_compact_fallback_started":
            latest_question = "整卷"
            latest_round = "紧凑high视觉定位兜底已发起"
            latest_event = name
        elif name == "answer_layout_scan_compact_fallback_finished":
            latest_question = "整卷"
            latest_round = f"紧凑high视觉定位完成，定位 {event.get('located_question_count', 0)} 题"
            latest_event = name
        elif name == "answer_layout_scan_compact_fallback_failed":
            latest_question = "整卷"
            latest_round = "紧凑high视觉定位失败，改用旧的页码启发式"
            latest_event = name
            latest_error = str(event.get("error") or "")
        elif name == "answer_layout_scan_failed":
            latest_question = "整卷"
            latest_round = "双xhigh视觉小格定位失败，改用旧的页码启发式"
            latest_event = name
            latest_error = str(event.get("error") or "")
        elif name in {"model_cache_hit"}:
            call_name = str(event.get("call_name") or "")
            match = re.search(r"q(\d+)_(?:round(\d+)|single)", call_name)
            latest_question = match.group(1) if match else ("整卷" if call_name == "answer_layout_scan" else "1-10" if call_name.startswith("objective_batch_") else latest_question)
            latest_round = "模型响应缓存命中"
            latest_event = name
        elif name in {"model_call_started", "model_call", "model_call_error"}:
            call_name = str(event.get("call_name") or "")
            if call_name:
                model_tasks.add(call_name)
            attempt_no = as_int(event.get("attempt_no"), 1)
            retry_no = as_int(event.get("retry_no"), 1)
            if attempt_no > 1 or retry_no > 1:
                retry_or_degrade_calls += 1
            if name == "model_call_started":
                active_calls += 1
            else:
                if name == "model_call" and as_int(event.get("status_code")) < 400:
                    successful_calls += 1
                else:
                    failed_calls += 1
                active_calls = max(0, active_calls - 1)
            match = re.search(r"q(\d+)_(?:round(\d+)|single)", call_name)
            if match:
                latest_question = match.group(1)
                round_no = match.group(2)
                latest_round = f"第 {round_no} 轮模型评分" if round_no else "单评模型评分"
                if name == "model_call_started":
                    latest_round += "已发起"
            elif call_name.startswith("objective_batch_"):
                latest_question = "1-10"
                if call_name == "objective_batch_single":
                    latest_round = "选择题视觉读选项并本地判分"
                else:
                    round_no = call_name.replace("objective_batch_round", "")
                    latest_round = f"选择题视觉读选项第 {round_no} 轮"
                if name == "model_call_started":
                    latest_round += "已发起"
            elif call_name.startswith("answer_layout_scan"):
                latest_question = "整卷"
                latest_round = "双xhigh视觉小格定位"
                if name == "model_call_started":
                    latest_round += "已发起"
            latest_event = name
            if event.get("elapsed_seconds") is not None:
                latest_elapsed = event.get("elapsed_seconds")
            if event.get("header_elapsed_seconds") is not None and event.get("elapsed_seconds") is not None:
                latest_round = (latest_round or "模型调用") + f"（首包 {event.get('header_elapsed_seconds')}s）"
            if event.get("error"):
                latest_error = str(event.get("error"))
    text = "等待任务开始"
    if latest_question:
        text = f"最近处理第 {latest_question} 题"
        if latest_round:
            text += f"，{latest_round}"
        if latest_elapsed is not None:
            text += f"，上次调用 {latest_elapsed}s"
    if latest_error:
        text += f"，最近错误：{latest_error[:120]}"
    if model_tasks:
        text += f"，模型任务 {len(model_tasks)} 个"
    if successful_calls:
        text += f"，成功 {successful_calls} 次"
    if failed_calls:
        text += f"，失败/重试 {failed_calls} 次"
    if retry_or_degrade_calls:
        text += f"，额外重试/降级 {retry_or_degrade_calls} 次"
    if active_calls:
        text += f"，活跃调用约 {active_calls} 次"
    return {
        "latest_question": latest_question,
        "latest_round": latest_round,
        "latest_event": latest_event,
        "layout_scan_status": layout_scan_status,
        "latest_elapsed_seconds": latest_elapsed,
        "latest_error": latest_error,
        "completed_model_calls": successful_calls,
        "successful_model_calls": successful_calls,
        "failed_model_calls": failed_calls,
        "model_task_count": len(model_tasks),
        "retry_or_degrade_model_calls": retry_or_degrade_calls,
        "active_model_calls": active_calls,
        "text": text,
    }


def latest_runs() -> list[dict[str, str]]:
    if not OUTPUT_DIR.exists():
        return []
    rows: list[dict[str, object]] = []
    child_runs_by_parent: dict[str, list[Path]] = {}
    ignored = {"web_server", "cache", "question_result_cache", "scorebooks"}
    for path in OUTPUT_DIR.iterdir():
        if not path.is_dir():
            continue
        if path.name in ignored:
            continue
        match = re.match(r"^(web_\d{8}_\d{6}_[A-Za-z0-9]{6})_\d{3}_", path.name)
        if match:
            child_runs_by_parent.setdefault(match.group(1), []).append(path)

    for parent_id, child_paths in child_runs_by_parent.items():
        child_paths = sorted(child_paths, key=lambda item: item.name)
        report_count = sum(1 for child in child_paths if (child / "grading_report.json").exists())
        mtime = max(child.stat().st_mtime for child in child_paths)
        rows.append(
            {
                "run_id": parent_id,
                "path": str(OUTPUT_DIR),
                "mtime": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(mtime)),
                "has_report": report_count > 0,
                "is_batch": True,
                "child_count": len(child_paths),
                "report_count": report_count,
                "child_run_ids": [child.name for child in child_paths],
                "view_run_id": child_paths[0].name if child_paths else parent_id,
            }
        )

    for path in OUTPUT_DIR.iterdir():
        if not path.is_dir():
            continue
        if path.name in ignored:
            continue
        if any(path in child_paths for child_paths in child_runs_by_parent.values()):
            continue
        report = path / "grading_report.json"
        rows.append(
            {
                "run_id": path.name,
                "path": str(path),
                "mtime": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(path.stat().st_mtime)),
                "has_report": report.exists(),
                "is_batch": False,
                "child_count": 1,
                "report_count": 1 if report.exists() else 0,
                "child_run_ids": [],
                "view_run_id": path.name,
            }
        )
    return sorted(rows, key=lambda row: str(row.get("mtime", "")), reverse=True)[:20]


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt: str, *args: object) -> None:
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def do_GET(self) -> None:  # noqa: N802
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/":
                text_response(self, INDEX_HTML, content_type="text/html")
            elif parsed.path == "/api/files":
                json_response(
                    self,
                    {
                        "pdf_files": list_pdf_files(),
                        "root": str(ROOT),
                        "default_policy_path": str(DEFAULT_POLICY_PATH),
                        "policy_exists": DEFAULT_POLICY_PATH.exists(),
                    },
                )
            elif parsed.path == "/api/runs":
                json_response(self, {"runs": latest_runs(), "tasks": [task_to_dict(task) for task in TASKS.values()]})
            elif parsed.path == "/api/task":
                query = parse_qs(parsed.query)
                run_id = query.get("run_id", [""])[0]
                with TASK_LOCK:
                    task = TASKS.get(run_id)
                if not task:
                    json_response(self, {"found": False}, status=404)
                else:
                    json_response(self, {"found": True, "task": task_to_dict(task)})
            elif parsed.path == "/api/report":
                query = parse_qs(parsed.query)
                run_id = query.get("run_id", [""])[0]
                filename = query.get("file", ["grading_report.json"])[0]
                json_response(self, load_report(run_id, filename))
            elif parsed.path == "/api/scorebook":
                query = parse_qs(parsed.query)
                run_id = query.get("run_id", [""])[0]
                body, filename = build_scorebook(run_id)
                binary_response(
                    self,
                    body,
                    filename,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            elif parsed.path == "/api/audit":
                query = parse_qs(parsed.query)
                run_id = query.get("run_id", [""])[0]
                json_response(self, load_audit_tail(run_id))
            else:
                text_response(self, "Not found", status=404)
        except Exception as exc:  # noqa: BLE001
            json_response(self, {"error": f"{type(exc).__name__}: {exc}"}, status=500)

    def do_POST(self) -> None:  # noqa: N802
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/api/stop":
                payload = read_request_json(self)
                run_id = str(payload.get("run_id") or "").strip()
                state = stop_grading_task(run_id)
                json_response(self, {"ok": True, "task": task_to_dict(state)})
                return
            if parsed.path == "/api/upload":
                query = parse_qs(parsed.query)
                role = query.get("role", [""])[0]
                uploaded = save_uploaded_pdfs(self, role)
                json_response(self, {"ok": True, "files": uploaded, "file": uploaded[0], "pdf_files": list_pdf_files()})
                return
            if parsed.path == "/api/delete-files":
                payload = read_request_json(self)
                raw_paths = payload.get("paths") or []
                if not isinstance(raw_paths, list):
                    raise ValueError("paths must be a list")
                result = delete_uploaded_pdfs([str(item) for item in raw_paths])
                json_response(self, {"ok": True, **result})
                return
            if parsed.path != "/api/grade":
                text_response(self, "Not found", status=404)
                return
            payload = read_request_json(self)
            state = start_grading_task(payload)
            json_response(self, {"ok": True, "task": task_to_dict(state)})
        except Exception as exc:  # noqa: BLE001
            json_response(self, {"ok": False, "error": f"{type(exc).__name__}: {exc}"}, status=400)


def task_to_dict(task: TaskState) -> dict:
    return {
        "run_id": task.run_id,
        "status": task.status,
        "created_at": task.created_at,
        "started_at": task.started_at,
        "finished_at": task.finished_at,
        "returncode": task.returncode,
        "command_preview": task.command_preview,
        "stdout": task.stdout,
        "stderr": task.stderr,
        "output_dir": task.output_dir,
        "error": task.error,
        "batch_index": task.batch_index,
        "batch_total": task.batch_total,
        "child_run_ids": task.child_run_ids,
        "current_child_run_id": task.current_child_run_id,
        "batch_results": task.batch_results,
    }


INDEX_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>考研数学自动阅卷本地原型</title>
  <style>
    :root {
      color-scheme: light;
      --ink: #111827;
      --muted: #667085;
      --line: #d0d7de;
      --panel: #ffffff;
      --panel-soft: #f8fafc;
      --bg: #f3f5f8;
      --accent: #2563eb;
      --accent-dark: #1d4ed8;
      --teal: #0f766e;
      --danger: #b42318;
      --warn: #b45309;
      --ok: #15803d;
      --purple: #6d28d9;
      --shadow: 0 12px 28px rgba(15, 23, 42, .08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif;
      background: var(--bg);
      color: var(--ink);
      font-size: 14px;
      line-height: 1.45;
    }
    header {
      background: #ffffff;
      border-bottom: 1px solid var(--line);
      padding: 14px 24px;
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 16px;
      position: sticky;
      top: 0;
      z-index: 20;
    }
    h1 { margin: 0; font-size: 19px; letter-spacing: 0; }
    h2 { margin: 0 0 12px; font-size: 15px; }
    .brand { display: flex; flex-direction: column; gap: 3px; min-width: 0; }
    .brand .hint { margin-top: 0; }
    main {
      padding: 18px 24px 32px;
      display: grid;
      grid-template-columns: minmax(380px, 470px) minmax(560px, 1fr);
      gap: 18px;
      align-items: start;
    }
    section {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      box-shadow: var(--shadow);
    }
    .control-panel {
      position: sticky;
      top: 76px;
      max-height: calc(100vh - 92px);
      overflow: auto;
    }
    .result-panel {
      min-width: 0;
    }
    label { display: block; font-weight: 600; margin: 12px 0 6px; }
    input, select, textarea, button {
      font: inherit;
    }
    input, select {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 9px 10px;
      background: #fff;
      min-height: 38px;
    }
    input:focus, select:focus {
      outline: 2px solid rgba(37, 99, 235, .18);
      border-color: var(--accent);
    }
    .file-row {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 8px;
      align-items: center;
    }
    .file-manager {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      margin-top: 8px;
      background: var(--panel-soft);
    }
    .file-manager select {
      min-height: 90px;
      background: #fff;
    }
    .file-manager.single select {
      min-height: 78px;
    }
    .file-tools {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 8px;
    }
    .file-tools button {
      padding: 7px 10px;
      min-height: 32px;
    }
    .file-meta {
      display: flex;
      justify-content: space-between;
      gap: 8px;
      color: var(--muted);
      font-size: 12px;
      margin-top: 6px;
    }
    input[type="file"] {
      padding: 8px;
      min-height: 38px;
    }
    .row { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
    .row.triple { grid-template-columns: repeat(3, minmax(0, 1fr)); }
    .row.triple label { margin-top: 0; }
    .checks { display: flex; flex-wrap: wrap; gap: 12px; margin-top: 12px; }
    .checks label { margin: 0; display: flex; align-items: center; gap: 6px; font-weight: 500; }
    .checks input { width: auto; min-height: 0; }
    button {
      border: 0;
      border-radius: 6px;
      padding: 10px 14px;
      background: var(--accent);
      color: #fff;
      cursor: pointer;
      min-height: 38px;
    }
    button:hover { background: var(--accent-dark); }
    button.secondary { background: #edf2f7; color: var(--ink); border: 1px solid #d7dee8; }
    button.secondary:hover { background: #e2e8f0; }
    button:disabled { opacity: .55; cursor: not-allowed; }
    .actions { display: flex; gap: 10px; margin-top: 16px; }
    button.danger { background: var(--danger); }
    button.danger:hover { background: #8f1d14; }
    .hint { color: var(--muted); font-size: 12px; margin-top: 6px; }
    .status {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 5px 10px;
      background: #fff;
      color: var(--muted);
      white-space: nowrap;
    }
    .status.running { color: var(--warn); border-color: #f1c56b; background: #fff8e5; }
    .status.finished { color: var(--ok); border-color: #95d5b2; background: #effaf4; }
    .status.failed { color: var(--danger); border-color: #f0a3a0; background: #fff1f0; }
    .toolbar {
      display: flex;
      gap: 10px;
      align-items: center;
      justify-content: space-between;
      margin-bottom: 12px;
    }
    .result-head {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 12px;
    }
    .result-actions {
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      justify-content: flex-end;
    }
    .run-path {
      margin-top: 4px;
      max-width: 760px;
      overflow-wrap: anywhere;
    }
    .report-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(90px, 1fr));
      gap: 10px;
      margin-bottom: 12px;
    }
    .metric {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      background: var(--panel-soft);
    }
    .metric strong { display: block; font-size: 20px; line-height: 1.2; }
    .metric span { color: var(--muted); font-size: 12px; }
    .metric:nth-child(1) { border-top: 3px solid var(--accent); }
    .metric:nth-child(2) { border-top: 3px solid var(--teal); }
    .metric:nth-child(3) { border-top: 3px solid var(--warn); }
    .metric:nth-child(4) { border-top: 3px solid var(--purple); }
    .progress-box {
      border: 1px solid #d6e0ea;
      border-left: 4px solid var(--accent);
      border-radius: 8px;
      padding: 10px 12px;
      background: #f8fbff;
      color: #334155;
      margin-bottom: 10px;
    }
    table {
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      margin-top: 8px;
      background: #fff;
    }
    th, td {
      border-bottom: 1px solid var(--line);
      padding: 8px;
      text-align: left;
      vertical-align: top;
      word-break: break-word;
    }
    th { background: #f8fafc; font-weight: 700; }
    thead th {
      position: sticky;
      top: 63px;
      z-index: 5;
      box-shadow: inset 0 -1px 0 var(--line);
    }
    tbody tr:hover { background: #fbfdff; }
    pre {
      margin: 0;
      padding: 12px;
      background: #111827;
      color: #eef2ff;
      border-radius: 8px;
      overflow: auto;
      max-height: 360px;
      white-space: pre-wrap;
      word-break: break-word;
    }
    .tabs { display: flex; gap: 8px; margin: 14px 0 10px; border-bottom: 1px solid var(--line); }
    .tabs button { background: #e7eef4; color: var(--ink); }
    .tabs button.active { background: var(--accent); color: white; border-color: var(--accent); }
    .run-list {
      margin-top: 14px;
      border-top: 1px solid var(--line);
      padding-top: 12px;
    }
    .run-item {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 8px;
      padding: 6px 0;
      border-bottom: 1px solid #edf0f4;
    }
    .run-actions { display: flex; gap: 6px; flex-wrap: wrap; justify-content: flex-end; }
    .run-item button { padding: 6px 9px; min-height: 30px; }
    @media (max-width: 980px) {
      main { grid-template-columns: 1fr; padding: 14px; }
      .report-grid { grid-template-columns: repeat(2, 1fr); }
      header { align-items: flex-start; flex-direction: column; }
      .control-panel { position: static; max-height: none; }
      .result-head { flex-direction: column; }
      thead th { top: 0; }
    }
  </style>
</head>
<body>
  <header>
    <div class="brand">
      <h1>考研数学自动阅卷本地原型</h1>
      <div class="hint" id="rootHint">正在读取本地文件...</div>
    </div>
    <div class="status" id="statusBadge">未运行</div>
  </header>
  <main>
    <section class="control-panel">
      <h2>阅卷配置</h2>
      <label for="submissionPdf">学生答卷 PDF</label>
      <div class="file-row">
        <input id="submissionUpload" type="file" accept="application/pdf,.pdf" multiple />
        <button class="secondary" data-upload-role="submission">批量导入</button>
      </div>
      <div class="file-manager">
        <select id="submissionPdf" multiple size="6"></select>
        <div class="file-meta">
          <span id="submissionCount">尚未导入学生答卷</span>
          <span>多选后按顺序批改</span>
        </div>
        <div class="file-tools">
          <button class="secondary" data-select-all="submissionPdf">全选学生卷</button>
          <button class="secondary" data-clear-select="submissionPdf">清空选择</button>
          <button class="danger" data-delete-select="submissionPdf">删除所选</button>
        </div>
      </div>
      <label for="questionPaperPdf">试卷题目 PDF</label>
      <div class="file-row">
        <input id="questionPaperUpload" type="file" accept="application/pdf,.pdf" />
        <button class="secondary" data-upload-role="question_paper">导入</button>
      </div>
      <div class="file-manager single">
        <select id="questionPaperPdf" size="3"></select>
        <div class="file-meta">
          <span id="questionPaperCount">尚未导入试卷题目</span>
          <span>当前选中即为本次题目 PDF</span>
        </div>
        <div class="file-tools">
          <button class="danger" data-delete-current="questionPaperPdf">删除当前题目</button>
        </div>
      </div>
      <label for="referencePdf">参考答案 / 评分依据 PDF</label>
      <div class="file-row">
        <input id="referenceUpload" type="file" accept="application/pdf,.pdf" />
        <button class="secondary" data-upload-role="reference">导入</button>
      </div>
      <div class="file-manager single">
        <select id="referencePdf" size="3"></select>
        <div class="file-meta">
          <span id="referenceCount">尚未导入参考答案</span>
          <span>当前选中即为本次评分依据</span>
        </div>
        <div class="file-tools">
          <button class="danger" data-delete-current="referencePdf">删除当前答案</button>
        </div>
      </div>
      <div class="hint" id="uploadHint">学生答卷可一次多选批量导入；删除按钮只删除项目临时副本，不会删除微信或磁盘中的原始 PDF。</div>
      <div class="row">
        <div>
          <label for="questions">题号</label>
          <input id="questions" value="1-22" placeholder="例如 1-10,11-16,17-22" />
        </div>
        <div>
          <label for="model">模型</label>
          <input id="model" value="gpt-5.5" />
        </div>
      </div>
      <label for="apiUrl">API URL</label>
      <input id="apiUrl" value="" placeholder="例如 https://your-gateway.example.com/v1" />
      <label for="userAgent">请求头 User-Agent</label>
      <input id="userAgent" value="Codex Desktop/0.133.0-alpha.1 (Windows 10.0.19045; x86_64) unknown (Codex Desktop; 9.41501)" />
      <label for="apiMode">调用方式</label>
      <select id="apiMode">
        <option value="responses" selected>/v1/responses</option>
        <option value="chat">/v1/chat/completions</option>
      </select>
      <div class="hint">Responses 作为默认主通道；解答题会结合自动解析文本与原始 PDF 页图进行视觉复核。</div>
      <label>分题型推理强度</label>
      <div class="row triple">
        <div>
          <label for="objectiveReasoning">选择题</label>
          <select id="objectiveReasoning">
            <option value="high" selected>high</option>
            <option value="xhigh">xhigh</option>
            <option value="medium">medium</option>
            <option value="low">low</option>
          </select>
        </div>
        <div>
          <label for="blankReasoning">填空题</label>
          <select id="blankReasoning">
            <option value="high" selected>high</option>
            <option value="xhigh">xhigh</option>
            <option value="medium">medium</option>
            <option value="low">low</option>
          </select>
        </div>
        <div>
          <label for="solutionReasoning">解答题</label>
          <select id="solutionReasoning">
            <option value="high" selected>high</option>
            <option value="xhigh">xhigh</option>
            <option value="medium">medium</option>
            <option value="low">low</option>
          </select>
        </div>
      </div>
      <div class="hint">默认先用 2 次 xhigh 视觉并行总览，切出选择题答案串、填空小格和大题作答框；后续每题用 high 结合 MinerU OCR 与局部视觉图阅卷。</div>
      <label for="apiKey">API Key</label>
      <input id="apiKey" type="password" placeholder="只在本地进程中用于本次调用，不写入文件" />
      <label for="policyPath">全局阅卷规范</label>
      <input id="policyPath" value="" />
      <div class="hint">默认使用项目内蒸馏的考研数学阅卷规范，作为每次评分的固定上下文。</div>
      <div class="row">
        <div>
          <label for="paperId">试卷 ID</label>
          <input id="paperId" value="2026-05-fudan-math" />
        </div>
        <div>
          <label for="candidate">考生名</label>
          <input id="candidate" value="本地测试" />
        </div>
      </div>
      <div class="row">
        <div>
          <label for="timeout">单次超时秒数</label>
          <input id="timeout" type="number" value="300" min="20" max="900" />
        </div>
        <div>
          <label for="retries">重试次数</label>
          <input id="retries" type="number" value="3" min="1" max="6" />
        </div>
      </div>
      <label for="maxOutputTokens">单次最大输出 tokens</label>
      <input id="maxOutputTokens" type="number" value="100000" min="200" max="200000" />
      <label for="concurrency">题目并发数</label>
      <input id="concurrency" type="number" value="5" min="1" max="10" />
      <div class="row">
        <div>
          <label for="blankConcurrency">填空并发</label>
          <input id="blankConcurrency" type="number" value="2" min="1" max="6" />
        </div>
        <div>
          <label for="solutionConcurrency">大题并发</label>
          <input id="solutionConcurrency" type="number" value="3" min="1" max="6" />
        </div>
      </div>
      <div class="hint">当前为稳态并发：选择题批任务 1 路，填空最多 2 路，大题最多 3 路；中转站不稳时比 10 路更容易跑完整张卷。</div>
      <div class="checks">
        <label><input id="parseOnly" type="checkbox" /> 只解析不调用模型</label>
        <label><input id="useCache" type="checkbox" checked /> 调试缓存</label>
        <label><input id="useStream" type="checkbox" checked disabled /> 流式接收</label>
        <label><input id="layoutScan" type="checkbox" checked /> 先整卷视觉定位</label>
        <label><input id="objectiveBatchMode" type="checkbox" checked /> 选择题10题批量</label>
        <label><input id="singleReview" type="checkbox" checked /> 单评极速</label>
        <label><input id="parallelVisualRounds" type="checkbox" /> 双评时同题两轮并发</label>
        <label><input id="refOfficial" type="checkbox" /> 参考材料标为官方</label>
        <label><input id="strictOfficial" type="checkbox" /> 严格官方依据</label>
      </div>
      <div class="actions">
        <button id="startBtn">开始阅卷</button>
        <button class="danger" id="stopBtn" disabled>停止任务</button>
        <button class="secondary" id="refreshBtn">刷新文件</button>
      </div>
      <div class="hint">全卷默认 1-22；先做整卷视觉导航，再按选择题批任务、填空题、大题三组并发推进。</div>
      <div class="run-list">
        <h2>最近结果</h2>
        <div id="runs"></div>
      </div>
    </section>
    <section class="result-panel">
      <div class="result-head">
        <div>
          <h2>运行结果</h2>
          <div id="runId" class="hint run-path"></div>
        </div>
        <div class="result-actions">
          <button id="downloadScorebookBtn" class="secondary" disabled>下载得分表</button>
          <button id="refreshReportBtn" class="secondary" disabled>刷新报告</button>
        </div>
      </div>
      <div class="report-grid">
        <div class="metric"><strong id="totalScore">-</strong><span>总分</span></div>
        <div class="metric"><strong id="gradedCount">-</strong><span>已评题数</span></div>
        <div class="metric"><strong id="reviewCount">-</strong><span>需复核</span></div>
        <div class="metric"><strong id="arbCount">-</strong><span>仲裁次数</span></div>
      </div>
      <div id="progressBox" class="progress-box">双xhigh视觉小格定位：等待开始。等待任务开始</div>
      <div id="summary"></div>
      <div class="tabs">
        <button class="active" data-tab="table">得分表</button>
        <button data-tab="json">报告 JSON</button>
        <button data-tab="audit">日志摘要</button>
      </div>
      <div class="hint">得分表也可导出为 Excel：单张卷为一个 sheet，批量任务为同一工作簿多个 sheet。</div>
      <div id="tablePanel"></div>
      <pre id="jsonPanel" style="display:none"></pre>
      <pre id="auditPanel" style="display:none"></pre>
    </section>
  </main>
  <script>
    let currentRunId = null;
    let scorebookRunId = null;
    let displayedRunId = null;
    let pollTimer = null;
    let activeTab = "table";

    const $ = (id) => document.getElementById(id);

    function setStatus(text, cls) {
      const el = $("statusBadge");
      el.textContent = text;
      el.className = "status" + (cls ? " " + cls : "");
    }

    async function api(path, options) {
      const res = await fetch(path, options);
      const data = await res.json();
      if (!res.ok) throw new Error(data.error || res.statusText);
      return data;
    }

    function selectValue(select, value) {
      if (!value) return;
      for (const option of select.options) {
        if (option.value === value) {
          select.value = value;
          return;
        }
      }
    }

    function selectValues(select, values) {
      const wanted = new Set((values || []).filter(Boolean));
      for (const option of select.options) {
        option.selected = wanted.has(option.value);
      }
    }

    function selectedValues(select) {
      return Array.from(select.selectedOptions || []).map(option => option.value).filter(Boolean);
    }

    function setScorebookRun(runId, enabled = true) {
      scorebookRunId = runId || null;
      const btn = $("downloadScorebookBtn");
      btn.disabled = !enabled || !scorebookRunId;
      btn.textContent = scorebookRunId ? "下载得分表" : "等待得分表";
    }

    async function loadFiles() {
      const data = await api("/api/files");
      $("rootHint").textContent = "项目目录：" + data.root;
      $("policyPath").value = data.default_policy_path || "";
      const files = data.pdf_files || [];
      fillRoleSelect($("submissionPdf"), files, "submission", "2702124-数学-复旦大学.pdf", "submissionCount", "学生答卷");
      fillRoleSelect($("questionPaperPdf"), files, "question_paper", "五月数学模考(3).pdf", "questionPaperCount", "试卷题目");
      fillRoleSelect($("referencePdf"), files, "reference", "五月数学模考答案(1).pdf", "referenceCount", "参考答案");
      await loadRuns();
    }

    function fillRoleSelect(select, files, role, preferredName, countId, label) {
      select.dataset.countId = countId;
      select.dataset.fileLabel = label;
      const oldValues = selectedValues(select);
      const old = select.value;
      select.innerHTML = "";
      const roleFiles = files.filter(file => file.role === role);
      for (const file of roleFiles) {
        const option = document.createElement("option");
        option.value = file.path;
        option.textContent = `${file.name} · ${formatSize(file.size_bytes)} · ${file.mtime || ""}`;
        option.dataset.deletable = file.deletable ? "1" : "0";
        option.dataset.group = file.group || "";
        select.appendChild(option);
        if (file.name === preferredName) option.selected = true;
      }
      if (select.multiple && oldValues.length) {
        selectValues(select, oldValues);
      } else if (old) {
        select.value = old;
      }
      if (!select.multiple && !select.value && select.options.length) {
        select.selectedIndex = 0;
      }
      const selectedCount = selectedValues(select).length;
      $(countId).textContent = select.multiple
        ? `${label} ${roleFiles.length} 份，已选 ${selectedCount} 份`
        : `${label} ${roleFiles.length} 份${select.value ? "，已选 1 份" : ""}`;
      if (!roleFiles.length) {
        const option = document.createElement("option");
        option.value = "";
        option.textContent = `尚未导入${label} PDF`;
        option.disabled = true;
        select.appendChild(option);
      }
      refreshFileCount(select);
    }

    function refreshFileCount(select) {
      const countId = select.dataset.countId;
      if (!countId) return;
      const label = select.dataset.fileLabel || "PDF";
      const total = Array.from(select.options).filter(option => option.value).length;
      const selectedCount = selectedValues(select).length;
      if (!total) {
        $(countId).textContent = `尚未导入${label}`;
      } else if (select.multiple) {
        $(countId).textContent = `${label} ${total} 份，已选 ${selectedCount} 份`;
      } else {
        $(countId).textContent = `${label} ${total} 份${selectedCount ? "，已选 1 份" : ""}`;
      }
    }

    function formatSize(bytes) {
      const value = Number(bytes || 0);
      if (value >= 1024 * 1024) return `${(value / 1024 / 1024).toFixed(1)}MB`;
      if (value >= 1024) return `${Math.round(value / 1024)}KB`;
      return `${value}B`;
    }

    async function uploadPdf(role) {
      const map = {
        submission: {input: "submissionUpload", select: "submissionPdf", label: "学生答卷"},
        question_paper: {input: "questionPaperUpload", select: "questionPaperPdf", label: "试卷题目"},
        reference: {input: "referenceUpload", select: "referencePdf", label: "参考答案"}
      };
      const cfg = map[role];
      const fileInput = $(cfg.input);
      const files = Array.from(fileInput.files || []);
      if (!files.length) throw new Error(`请先选择${cfg.label} PDF`);
      for (const file of files) {
        if (!file.name.toLowerCase().endsWith(".pdf")) throw new Error("只能导入 PDF 文件");
      }
      $("uploadHint").textContent = `正在导入 ${files.length} 个 PDF ...`;
      const form = new FormData();
      for (const file of files) form.append("file", file, file.name);
      const data = await api("/api/upload?role=" + encodeURIComponent(role), {
        method: "POST",
        body: form
      });
      await loadFiles();
      const uploadedFiles = data.files || (data.file ? [data.file] : []);
      if (role === "submission") {
        selectValues($(cfg.select), uploadedFiles.map(file => file.path));
        refreshFileCount($(cfg.select));
      } else if (uploadedFiles[0]) {
        selectValue($(cfg.select), uploadedFiles[0].path);
        refreshFileCount($(cfg.select));
      }
      $("uploadHint").textContent = `已导入 ${uploadedFiles.length} 个 PDF，可以继续选择其他 PDF 或开始阅卷。`;
      fileInput.value = "";
    }

    async function deleteSelectedFiles(selectId, onlyCurrent) {
      const select = $(selectId);
      const paths = onlyCurrent ? [select.value].filter(Boolean) : selectedValues(select);
      if (!paths.length) throw new Error("请先选择要删除的 PDF");
      const label = select.dataset.fileLabel || "PDF";
      const message = `确定删除 ${paths.length} 份${label}的项目临时副本吗？\n\n不会删除微信或磁盘中的原始文件。`;
      if (!confirm(message)) return;
      const data = await api("/api/delete-files", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify({paths})
      });
      await loadFiles();
      const deleted = data.deleted || [];
      const skipped = data.skipped || [];
      $("uploadHint").textContent = `已删除 ${deleted.length} 份临时 PDF${skipped.length ? `，${skipped.length} 份未删除：${skipped[0].reason}` : "。"} `;
    }

    async function loadRuns() {
      const data = await api("/api/runs");
      const box = $("runs");
      box.innerHTML = "";
      const taskRuns = [];
      for (const task of data.tasks || []) {
        if ((task.child_run_ids || []).length) {
          for (const row of task.batch_results || []) {
            taskRuns.push({
              run_id: row.run_id,
              mtime: statusText(row.status),
              has_report: row.status === "finished",
              path: row.output_dir
            });
          }
        }
      }
      for (const run of [...taskRuns, ...data.runs].slice(0, 8)) {
        const row = document.createElement("div");
        row.className = "run-item";
        const batchMark = run.is_batch ? `批量 ${run.report_count || 0}/${run.child_count || 0}` : (run.has_report ? "有报告" : "无完整报告");
        row.innerHTML = `<div><strong>${escapeHtml(run.run_id)}</strong><div class="hint">${escapeHtml(run.mtime)} ${escapeHtml(batchMark)}</div></div>`;
        const actions = document.createElement("div");
        actions.className = "run-actions";
        const btn = document.createElement("button");
        btn.className = "secondary";
        btn.textContent = "查看";
        btn.onclick = () => showRun(run.view_run_id || run.run_id, run.run_id);
        actions.appendChild(btn);
        const exportBtn = document.createElement("button");
        exportBtn.className = "secondary";
        exportBtn.textContent = "导出";
        exportBtn.disabled = !run.has_report;
        exportBtn.onclick = () => downloadScorebook(run.run_id);
        actions.appendChild(exportBtn);
        row.appendChild(actions);
        box.appendChild(row);
      }
    }

    async function startGrade() {
      const payload = {
        submission_pdfs: selectedValues($("submissionPdf")),
        submission_pdf: $("submissionPdf").value,
        question_paper_pdf: $("questionPaperPdf").value,
        reference_pdf: $("referencePdf").value,
        questions: $("questions").value,
        api_url: $("apiUrl").value,
        user_agent: $("userAgent").value,
        api_mode: $("apiMode").value,
        reasoning_effort: $("solutionReasoning").value,
        objective_reasoning_effort: $("objectiveReasoning").value,
        blank_reasoning_effort: $("blankReasoning").value,
        solution_reasoning_effort: $("solutionReasoning").value,
        model: $("model").value,
        api_key: $("apiKey").value,
        paper_id: $("paperId").value,
        candidate_name: $("candidate").value,
        timeout_seconds: Number($("timeout").value || 300),
        max_retries: Number($("retries").value || 3),
        max_output_tokens: Number($("maxOutputTokens").value || 100000),
        concurrency: Number($("concurrency").value || 10),
        blank_concurrency: Number($("blankConcurrency").value || 2),
        solution_concurrency: Number($("solutionConcurrency").value || 3),
        parse_only: $("parseOnly").checked,
        use_cache: $("useCache").checked,
        use_stream: true,
        layout_scan: $("layoutScan").checked,
        objective_batch_mode: $("objectiveBatchMode").checked,
        single_review: $("singleReview").checked,
        parallel_visual_rounds: $("parallelVisualRounds").checked,
        reference_is_official: $("refOfficial").checked,
        strict_official: $("strictOfficial").checked,
        policy_path: $("policyPath").value
      };
      $("startBtn").disabled = true;
      setStatus("运行中", "running");
      const data = await api("/api/grade", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify(payload)
      });
      currentRunId = data.task.run_id;
      setScorebookRun(currentRunId, false);
      $("runId").textContent = currentRunId;
      $("stopBtn").disabled = false;
      $("refreshReportBtn").disabled = false;
      pollTask();
      if (pollTimer) clearInterval(pollTimer);
      pollTimer = setInterval(pollTask, 2500);
    }

    async function pollTask() {
      if (!currentRunId) return;
      try {
        const data = await api("/api/task?run_id=" + encodeURIComponent(currentRunId));
        const task = data.task;
        setStatus(statusText(task.status), task.status);
        const visibleRunId = task.current_child_run_id || task.child_run_ids?.[Math.max(0, (task.batch_index || 1) - 1)] || task.run_id;
        const exportRunId = (task.status === "finished" && (task.child_run_ids || []).length) ? task.run_id : visibleRunId;
        const batchText = task.batch_total > 1 ? `批量 ${task.batch_index}/${task.batch_total} | ` : "";
        $("runId").textContent = `${batchText}${visibleRunId} | ${task.output_dir}`;
        setScorebookRun(exportRunId, task.status === "finished" || task.status === "failed" || Boolean(visibleRunId));
        await loadReportOrPartial(visibleRunId);
        if (task.status === "finished" || task.status === "failed") {
          $("startBtn").disabled = false;
          $("stopBtn").disabled = true;
          clearInterval(pollTimer);
          pollTimer = null;
          await loadRuns();
        }
      } catch (err) {
        setStatus("轮询失败", "failed");
        $("auditPanel").textContent = String(err);
        document.querySelector('[data-tab="audit"]').click();
      }
    }

    async function loadReportOrPartial(runId) {
      displayedRunId = runId;
      let report = await api(`/api/report?run_id=${encodeURIComponent(runId)}&file=grading_report.json`);
      if (!report.available && !report.score_summary) {
        report = await api(`/api/report?run_id=${encodeURIComponent(runId)}&file=partial_report.json`);
      }
      renderReport(report);
      const audit = await api(`/api/audit?run_id=${encodeURIComponent(runId)}`);
      $("auditPanel").textContent = JSON.stringify(audit, null, 2);
      renderProgress(report, audit);
      setScorebookRun(scorebookRunId || runId, Boolean((report.question_scores || []).length || report.score_summary));
      return report;
    }

    async function showRun(runId, exportRunId = null) {
      currentRunId = exportRunId || runId;
      setScorebookRun(exportRunId || runId, true);
      $("runId").textContent = exportRunId && exportRunId !== runId ? `${exportRunId} | 查看 ${runId}` : runId;
      setStatus("查看结果", "finished");
      $("stopBtn").disabled = true;
      $("refreshReportBtn").disabled = false;
      await loadReportOrPartial(runId);
    }

    async function stopTask() {
      if (!currentRunId) return;
      $("stopBtn").disabled = true;
      setStatus("停止中", "running");
      const data = await api("/api/stop", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify({run_id: currentRunId})
      });
      setStatus(statusText(data.task.status), data.task.status);
      await loadReportOrPartial(currentRunId);
      $("startBtn").disabled = false;
    }

    async function downloadScorebook(runId) {
      const targetRunId = runId || scorebookRunId || currentRunId;
      if (!targetRunId) return;
      const btn = $("downloadScorebookBtn");
      const oldText = btn.textContent;
      btn.disabled = true;
      btn.textContent = "正在生成...";
      try {
        const res = await fetch(`/api/scorebook?run_id=${encodeURIComponent(targetRunId)}`);
        if (!res.ok) {
          let message = res.statusText;
          try {
            const data = await res.json();
            message = data.error || message;
          } catch (_) {}
          throw new Error(message);
        }
        const blob = await res.blob();
        const disposition = res.headers.get("Content-Disposition") || "";
        const filename = scorebookFilename(disposition, targetRunId);
        const url = URL.createObjectURL(blob);
        const link = document.createElement("a");
        link.href = url;
        link.download = filename;
        document.body.appendChild(link);
        link.click();
        link.remove();
        URL.revokeObjectURL(url);
      } catch (err) {
        alert("得分表导出失败：" + err);
      } finally {
        btn.disabled = !scorebookRunId && !currentRunId;
        btn.textContent = oldText;
      }
    }

    function scorebookFilename(disposition, runId) {
      const utf8Match = disposition.match(/filename\*=UTF-8''([^;]+)/i);
      if (utf8Match) {
        try { return decodeURIComponent(utf8Match[1]); } catch (_) {}
      }
      const asciiMatch = disposition.match(/filename="?([^";]+)"?/i);
      return asciiMatch ? asciiMatch[1] : `得分表_${runId}.xlsx`;
    }

    function renderReport(report) {
      $("jsonPanel").textContent = JSON.stringify(report, null, 2);
      const summary = report.score_summary || {};
      $("totalScore").textContent = formatScore(summary.total_score_for_graded_questions);
      const graded = summary.graded_question_count ?? report.completed_question_count ?? (report.question_scores || []).length ?? "-";
      const total = report.total_question_count ? ` / ${report.total_question_count}` : "";
      $("gradedCount").textContent = `${graded}${total}`;
      $("reviewCount").textContent = summary.human_review_count ?? (report.question_scores || []).filter(row => row.needs_human_review).length ?? "-";
      $("arbCount").textContent = report.audit_log_summary?.arbitration_count ?? "-";
      renderTable(report.question_scores || []);
    }

    function renderProgress(report, audit) {
      const completed = report.completed_question_count ?? report.score_summary?.graded_question_count ?? (report.question_scores || []).length ?? 0;
      const total = report.total_question_count ?? "";
      const text = audit.progress?.text || "等待任务开始";
      const layoutText = audit.progress?.layout_scan_status || "双xhigh视觉小格定位：等待开始";
      const countText = total ? `已完成 ${completed}/${total} 题。` : "";
      $("progressBox").textContent = `${layoutText}。${countText}${text}`;
    }

    function renderTable(rows) {
      if (!rows.length) {
        $("tablePanel").innerHTML = '<div class="hint">暂无得分表，任务运行中或尚未生成报告。</div>';
        return;
      }
      let html = `<table><thead><tr>
        <th style="width:58px">题号</th>
        <th style="width:80px">得分</th>
        <th style="width:90px">复核</th>
        <th style="width:220px">视觉复核作答</th>
        <th>得分点</th>
        <th>扣分点 / 原因</th>
      </tr></thead><tbody>`;
      for (const row of rows) {
        const score = row.final_score === null || row.final_score === undefined ? "待复核" : `${formatScore(row.final_score)} / ${formatScore(row.full_score)}`;
        const source = row.student_answer_source || {};
        const recognized = source.recognized_student_answer || row.recognized_student_answer || row.grading_round_1?.recognized_student_answer || row.grading_round_1?.student_choice || "";
        const visualSummary = source.visual_reading_summary || row.visual_reading_summary || row.grading_round_1?.visual_reading_summary || "";
        const evidenceUsed = source.evidence_used || row.evidence_used || row.grading_round_1?.evidence_used || "";
        const ocrDraft = source.student_answer_ocr || "";
        html += `<tr>
          <td>${escapeHtml(row.question_no)}</td>
          <td>${escapeHtml(score)}</td>
          <td>${row.needs_human_review ? "是" : "否"}${row.third_arbitration_triggered ? "<br>已仲裁" : ""}</td>
          <td>
            <div>${escapeHtml(compactCell(recognized || "未形成视觉复核文本"))}</div>
            <div class="hint">${escapeHtml(compactCell(visualSummary))}</div>
            <div class="hint">${escapeHtml(evidenceUsed ? "证据：" + evidenceUsed : "")}</div>
            <div class="hint">${escapeHtml(ocrDraft ? "MinerU草稿：" + compactCell(ocrDraft) : "")}</div>
          </td>
          <td>${listHtml(row.main_earned_points || row.valid_student_steps || [])}</td>
          <td>${listHtml(row.main_deducted_points || row.wrong_or_missing_steps || [])}<div class="hint">${escapeHtml(row.review_reason || "")}</div></td>
        </tr>`;
      }
      html += "</tbody></table>";
      $("tablePanel").innerHTML = html;
    }

    function listHtml(items) {
      if (!items.length) return '<span class="hint">无</span>';
      return items.slice(0, 4).map(item => `<div>${escapeHtml(item)}</div>`).join("");
    }

    function compactCell(value) {
      value = String(value || "").replace(/\s+/g, " ").trim();
      return value.length > 120 ? value.slice(0, 119) + "…" : value;
    }

    function formatScore(value) {
      if (value === null || value === undefined || value === "") return "-";
      const number = Number(value);
      if (!Number.isFinite(number)) return String(value);
      return Number.isInteger(number) ? String(number) : number.toFixed(1).replace(/\.0$/, "");
    }

    function escapeHtml(value) {
      return String(value ?? "").replace(/[&<>"']/g, ch => ({
        "&": "&amp;",
        "<": "&lt;",
        ">": "&gt;",
        '"': "&quot;",
        "'": "&#39;"
      }[ch]));
    }

    function statusText(status) {
      return {
        queued: "排队中",
        running: "运行中",
        stopping: "停止中",
        finished: "已完成",
        failed: "失败"
      }[status] || status || "未知";
    }

    document.querySelectorAll(".tabs button").forEach(btn => {
      btn.onclick = () => {
        activeTab = btn.dataset.tab;
        document.querySelectorAll(".tabs button").forEach(b => b.classList.toggle("active", b === btn));
        $("tablePanel").style.display = activeTab === "table" ? "" : "none";
        $("jsonPanel").style.display = activeTab === "json" ? "" : "none";
        $("auditPanel").style.display = activeTab === "audit" ? "" : "none";
      };
    });

    $("startBtn").onclick = () => startGrade().catch(err => {
      $("startBtn").disabled = false;
      $("stopBtn").disabled = true;
      setStatus("启动失败", "failed");
      $("auditPanel").textContent = String(err);
      document.querySelector('[data-tab="audit"]').click();
    });
    $("stopBtn").onclick = () => stopTask().catch(err => {
      $("stopBtn").disabled = false;
      setStatus("停止失败", "failed");
      $("auditPanel").textContent = String(err);
      document.querySelector('[data-tab="audit"]').click();
    });
    $("refreshBtn").onclick = () => loadFiles().catch(err => alert(err));
    $("downloadScorebookBtn").onclick = () => downloadScorebook(scorebookRunId || currentRunId);
    $("refreshReportBtn").onclick = () => {
      const runId = displayedRunId || currentRunId;
      if (runId) loadReportOrPartial(runId).catch(err => alert(err));
    };
    ["submissionPdf", "questionPaperPdf", "referencePdf"].forEach(id => {
      $(id).onchange = () => refreshFileCount($(id));
    });
    document.querySelectorAll("[data-select-all]").forEach(btn => {
      btn.onclick = () => {
        const select = $(btn.dataset.selectAll);
        for (const option of select.options) {
          if (option.value) option.selected = true;
        }
        refreshFileCount(select);
      };
    });
    document.querySelectorAll("[data-clear-select]").forEach(btn => {
      btn.onclick = () => {
        const select = $(btn.dataset.clearSelect);
        for (const option of select.options) option.selected = false;
        refreshFileCount(select);
      };
    });
    document.querySelectorAll("[data-delete-select]").forEach(btn => {
      btn.onclick = () => deleteSelectedFiles(btn.dataset.deleteSelect, false).catch(err => {
        $("uploadHint").textContent = String(err);
        alert(err);
      });
    });
    document.querySelectorAll("[data-delete-current]").forEach(btn => {
      btn.onclick = () => deleteSelectedFiles(btn.dataset.deleteCurrent, true).catch(err => {
        $("uploadHint").textContent = String(err);
        alert(err);
      });
    });
    document.querySelectorAll("[data-upload-role]").forEach(btn => {
      btn.onclick = () => uploadPdf(btn.dataset.uploadRole).catch(err => {
        $("uploadHint").textContent = String(err);
        alert(err);
      });
    });
    loadFiles().catch(err => {
      setStatus("初始化失败", "failed");
      $("auditPanel").textContent = String(err);
    });
  </script>
</body>
</html>
"""


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Local web UI for the automatic math grader prototype.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    url = f"http://{args.host}:{args.port}"
    print(f"Local grader UI: {url}", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server.", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
